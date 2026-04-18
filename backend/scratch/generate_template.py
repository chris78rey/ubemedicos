import openpyxl

def create_template():
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Especialidades"
    
    # Header
    sheet.cell(row=1, column=1, value="Nombre de Especialidad")
    
    # Example rows (optional, but requested as "empty template", maybe just header)
    # sheet.cell(row=2, column=1, value="Cardiología")
    
    wb.save("g:/codex_projects/ubemedicos/web_frontend/public/plantilla_especialidades.xlsx")
    print("Template created successfully.")

if __name__ == "__main__":
    create_template()
